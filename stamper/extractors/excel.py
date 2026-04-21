"""
excel.py - Excel spreadsheet extractor.
"""

from pathlib import Path
from .base import BaseExtractor


RULE_COLUMN_KEYWORDS = ["rule", "description", "requirement", "constraint", "policy"]
CONSTRAINT_COLUMN_KEYWORDS = ["must", "shall", "should", "may", "type", "strength", "level"]


class ExcelExtractor(BaseExtractor):
    """Extract content from Excel files."""

    def extract(self) -> dict:
        """
        Extract content from Excel spreadsheet.

        Requires: openpyxl package
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": "dependency_missing",
                "message": "openpyxl not installed. Run: pip install openpyxl",
                "sheets": [],
                "raw_text": ""
            }

        path = Path(self.content) if not isinstance(self.content, str) else Path(self.content)

        try:
            wb = load_workbook(str(path), data_only=True)
        except Exception as e:
            return {
                "error": "extraction_failed",
                "message": str(e),
                "sheets": [],
                "raw_text": ""
            }

        sheets = []
        all_text = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            headers = []

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                row_values = [str(cell) if cell is not None else "" for cell in row]

                # Skip empty rows
                if not any(cell.strip() for cell in row_values):
                    continue

                if i == 0:
                    headers = row_values
                else:
                    rows.append(dict(zip(headers, row_values)))

                all_text.extend(row_values)

            # Detect rule and constraint columns
            rule_col, constraint_col = self._detect_rule_columns(headers)

            sheets.append({
                "name": sheet_name,
                "headers": headers,
                "rows": rows,
                "rule_column": rule_col,
                "constraint_column": constraint_col
            })

        return {
            "frontmatter": {},
            "sheets": sheets,
            "sections": self._sheets_to_sections(sheets),
            "prohibitions_in_frontmatter": [],
            "tables": [],
            "raw_text": '\n'.join(filter(None, all_text))
        }

    def _detect_rule_columns(self, headers: list[str]) -> tuple:
        """Detect which columns contain rules and constraints."""
        rule_col = next(
            (h for h in headers if any(k in h.lower() for k in RULE_COLUMN_KEYWORDS)),
            headers[0] if headers else None
        )
        constraint_col = next(
            (h for h in headers if any(k in h.lower() for k in CONSTRAINT_COLUMN_KEYWORDS)),
            None
        )
        return rule_col, constraint_col

    def _sheets_to_sections(self, sheets: list[dict]) -> list[dict]:
        """Convert sheets to section format for unified processing."""
        sections = []
        for sheet in sheets:
            content_lines = []
            for row in sheet.get("rows", []):
                rule_col = sheet.get("rule_column")
                if rule_col and rule_col in row:
                    content_lines.append(row[rule_col])

            sections.append({
                "heading": sheet["name"],
                "level": 1,
                "content": '\n'.join(content_lines),
                "lists": [],
                "tables": []
            })
        return sections
